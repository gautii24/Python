import configparser
import openpyxl


class ReadFromFiles():

    @staticmethod
    def readFromProp(self, propertyKey):
        PathFile = "D:/FinalFrameWork/CredenceWeb_DataSource/MasterData.xlsx"
        self.config = configparser.RawConfigParser()
        self.config.read(PathFile)
        self.value = self.config.get('BROWSER', propertyKey)
        return self.value

    @staticmethod
    def readFromExcel(self, SheetName, rowNum, ColumnNum):
        excelPath = "D:/FinalFrameWork/CredenceWeb_DataSource/MasterData.xlsx"
        self.wb = openpyxl.load_workbook(excelPath)
        self.ws = self.wb.get_sheet_by_name(SheetName)
        self.value = self.ws.cell(row=rowNum, column=ColumnNum).value
        return self.value


    @staticmethod
    def countActiveRows(self, SheetName):
        excelPath = "D:/FinalFrameWork\CredenceWeb_DataSource/MasterData.xlsx"
        self.wb = openpyxl.load_workbook(excelPath)
        self.ws = self.wb.get_sheet_by_name(SheetName)
        self.maxRowCount = self.ws.max_row
        return self.maxRowCount

    

